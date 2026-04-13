#第一步先写读取 xlsx 文件的部分，目标是把账单的46行数据用 Python 读出来，跳过前17行说明文字，只保留第18行以后的真实数据.

# 1. 导入库
import openpyxl
import decimal
from pathlib import Path
from classifier import classify_by_rules
from classifier import classify_by_llm
# 6. 连接数据库
import pymysql

conn = pymysql.connect(
    host='127.0.0.1',
    user='sentinel',
    password='Sentinel@2026',
    database='pocket_sentinel',
    charset='utf8mb4'
)
# 2. 定义文件路径
XLSX_FILE = Path.home() / '微信支付账单流水文件(20260404-20260411)_20260411140900.xlsx'

# 3. 打开 xlsx，获取工作表
wb=openpyxl.load_workbook(XLSX_FILE)
ws=wb.active
# 4. 遍历第19行开始的数据，跳过说明行
for row in ws.iter_rows(min_row=19, values_only=True):
#    每一行的列顺序是：
#    [0]交易时间 [1]交易类型 [2]交易对方 [3]商品
#    [4]收/支    [5]金额     [6]支付方式  [7]当前状态
#    [8]交易单号 [9]商户单号 [10]备注

# 5. 对每一行做清洗
#    - 跳过空行
    if not row[0]:continue
#    - 金额转 Decimal
    amount = row[5]
    if isinstance(amount, str):
        amount = amount.replace(',', '')  # 去掉千分位逗号
        amount = decimal.Decimal(amount)  # 转成 Decimal
#    - 收/支 转成 0/1
    if row[4] == '收入':
        direction = 0
    elif row[4] == '支出':
        direction = 1
    print(row[2], amount, direction)    

# 7. 对每一行：
#    - 查 dim_merchants 有没有这个商户
    merchant=row[2]
    with conn.cursor() as cursor:
          cursor.execute("SELECT peer_id FROM dim_merchants WHERE original_name=%s", (merchant,))


#    - 没有就 INSERT 进去
          result = cursor.fetchone()
          if not result:
            cursor.execute("INSERT INTO dim_merchants (original_name) VALUES (%s)", (merchant,))

#    - 用交易单号前32位作为 trans_hash
          trans_hash = row[8][:32]
#    - INSERT IGNORE 写入 fact_transactions
          cursor.execute("""
            INSERT IGNORE INTO fact_transactions (
                trade_time, trade_type, amount, direction, payment_method, status, trans_hash, peer_id
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, (SELECT peer_id FROM dim_merchants WHERE original_name=%s))
            """,
            (row[0], row[1], amount, direction, row[6], row[7], trans_hash, merchant))
#插入商户之后，用 classify_by_rules(merchant) 拿到分类，把 category_id 更新到 dim_merchants 和 fact_transactions。
          category_info = classify_by_rules(merchant)
          if category_info:
             main_cat = category_info['main_cat']
          else:
             main_cat = classify_by_llm(merchant)

          if main_cat:
           cursor.execute("SELECT category_id FROM dim_categories WHERE main_cat=%s", (main_cat,))
           cat_row = cursor.fetchone()
           if cat_row:
            category_id = cat_row[0]
            cursor.execute("UPDATE dim_merchants SET category_id=%s WHERE original_name=%s", (category_id, merchant))
            cursor.execute("UPDATE fact_transactions SET category_id=%s WHERE trans_hash=%s", (category_id, trans_hash))
# 8. 打印完成信息：共处理 X 条
conn.commit()
print("完成！共处理 {} 条记录。".format(ws.max_row - 18))















